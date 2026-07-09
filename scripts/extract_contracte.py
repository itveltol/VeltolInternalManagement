#!/usr/bin/env python3
"""One-off extraction: Contracte2025-2026 sheet -> clean JSON for import.

Reads 'Program +Contracte.xlsx', sheet 'Contracte2025-2026', and produces
scripts/contracte_import.json with one object per row, shaped for the
projects import script (import_contracte.ts) to consume.
"""
import json
import re
from datetime import datetime

import openpyxl

SRC = "Program +Contracte.xlsx"
SHEET = "Contracte2025-2026"
OUT = "scripts/contracte_import.json"

COMPANY_MARKERS = ["SRL", "S.R.L", "S.C.", "SC ", "SA", "S.A", "PFA", " II"]

MANAGER_NAME_MAP = {
    "nagy ors": "Nagy Ors",
    "ors": "Nagy Ors",
    "vlad ciubuca": "Vlad Ciubuca",
    "ciubuca vlad": "Vlad Ciubuca",
    "szasz bela": "Szasz Bela",
    "szasz béla": "Szasz Bela",
    "szász béla": "Szasz Bela",
    "kendi róbert": "Kendi Róbert",
    "kendi robert": "Kendi Róbert",
    "andrei": "Andrei",
    "papp alpar": "Papp Alpar",
    "zoli": "Zoli",
}


def norm_spaces(s):
    return re.sub(r"\s+", " ", str(s)).strip()


def guess_client_type(name):
    upper = name.upper()
    if any(m in upper for m in COMPANY_MARKERS):
        return "company"
    return "person"


def parse_date_ddmmyyyy(v):
    if v is None:
        return None
    s = str(v).strip().rstrip(".")
    if not s or s == "-":
        return None
    for fmt in ("%d.%m.%Y", "%d.%m.%y"):
        try:
            return datetime.strptime(s, fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    return None


def parse_manager(raw):
    if raw is None:
        return None, None
    key = norm_spaces(raw).lower()
    mapped = MANAGER_NAME_MAP.get(key)
    if mapped:
        return mapped, None
    return None, norm_spaces(raw)


CAPACITY_SIMPLE_RE = re.compile(
    r"^\s*([0-9]+(?:[.,][0-9]+)?)\s*(kwp|kw|mw)?\s*\.?\s*$", re.IGNORECASE
)


def parse_capacity_kw_to_mw(raw):
    """Returns (mw_solar, unparsed_raw_text_or_None)."""
    if raw is None:
        return None, None
    if isinstance(raw, (int, float)):
        return round(raw / 1000.0, 6), None
    s = str(raw).strip()
    if not s or s == "-":
        return None, None
    m = CAPACITY_SIMPLE_RE.match(s.replace(",", "."))
    if not m:
        return None, s
    value = float(m.group(1))
    unit = (m.group(2) or "kw").lower()
    if unit == "mw":
        return round(value, 6), None
    return round(value / 1000.0, 6), None


def num(v):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return v
    s = str(v).strip().replace(",", ".")
    if not s or s == "-":
        return None
    try:
        return float(s)
    except ValueError:
        return None


def txt(v):
    if v is None:
        return None
    s = norm_spaces(v)
    if not s or s == "-":
        return None
    return s


SPARSE_LABELS = [
    ("Val. avans-RON", 9),
    ("Val. Livrare materiale-RON", 10),
    ("Val. la PIF-RON", 11),
    ("Montare nr. pers.", 14),
    ("Valoare Bonus", 15),
    ("Racord nr. pers.", 16),
    ("Estimare Durata Lucrare-Zile", 17),
    ("Data inceput lucrare", 18),
    ("Data Finalizarii", 19),
    ("Cheltuieli Combustibil", 20),
    ("Total Cheltuieli", 21),
    ("Profit", 22),
]


def main():
    wb = openpyxl.load_workbook(SRC, data_only=True)
    ws = wb[SHEET]

    rows = []
    stats = {
        "total": 0,
        "skipped_no_contractant": 0,
        "manager_matched": 0,
        "manager_unmatched": 0,
        "capacity_parsed": 0,
        "capacity_unparsed": 0,
        "value_eur_missing": 0,
    }

    for r in range(2, ws.max_row + 1):
        contractant = ws.cell(row=r, column=4).value
        nr_crt = ws.cell(row=r, column=3).value
        if contractant is None and nr_crt is None:
            continue

        stats["total"] += 1
        contractant_name = txt(contractant)
        if not contractant_name:
            stats["skipped_no_contractant"] += 1
            continue

        achitat = txt(ws.cell(row=r, column=1).value)
        contract_number = txt(nr_crt)
        value_eur_raw = num(ws.cell(row=r, column=7).value)
        # projects.value_eur is bigint; round to nearest whole euro
        value_eur = round(value_eur_raw) if value_eur_raw is not None else None
        if value_eur is None:
            stats["value_eur_missing"] += 1
        deadline = parse_date_ddmmyyyy(ws.cell(row=r, column=12).value)
        responsabil_raw = ws.cell(row=r, column=13).value
        manager_name, manager_unmatched = parse_manager(responsabil_raw)
        if manager_name:
            stats["manager_matched"] += 1
        elif manager_unmatched:
            stats["manager_unmatched"] += 1

        mw_solar, capacity_unparsed = parse_capacity_kw_to_mw(
            ws.cell(row=r, column=8).value
        )
        if mw_solar is not None:
            stats["capacity_parsed"] += 1
        elif capacity_unparsed:
            stats["capacity_unparsed"] += 1

        notes_parts = []
        info = txt(ws.cell(row=r, column=23).value)
        if info:
            notes_parts.append(info)
        if capacity_unparsed:
            notes_parts.append(f"Capacitate (raw): {capacity_unparsed}")
        if manager_unmatched:
            notes_parts.append(f"Responsabil: {manager_unmatched}")
        for label, col in SPARSE_LABELS:
            v = ws.cell(row=r, column=col).value
            v_txt = txt(v)
            if v_txt:
                notes_parts.append(f"{label}: {v_txt}")

        rows.append(
            {
                "row": r,
                "name": contractant_name,
                "client_name": contractant_name,
                "client_type": guess_client_type(contractant_name),
                "contract_number": contract_number,
                "value_eur": value_eur,
                "deadline": deadline,
                "manager_name": manager_name,
                "mw_solar": mw_solar,
                "paid_by": achitat,
                "notes": "\n".join(notes_parts) if notes_parts else None,
            }
        )

    with open(OUT, "w", encoding="utf-8") as f:
        json.dump(rows, f, ensure_ascii=False, indent=2)

    print(f"Wrote {len(rows)} rows to {OUT}")
    print(json.dumps(stats, indent=2))


if __name__ == "__main__":
    main()
